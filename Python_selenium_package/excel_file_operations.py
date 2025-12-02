import openpyxl


def read_username_data(rownum, columnnum):
    # load excel sheet with the help of openpyxl plugin
    # load_workbook is inactive by default
    # book.active is to activate the worksheet
    book = openpyxl.load_workbook("C:\\Users\\jishu\\pythonProject\\excel_worksheet_python.xlsx")
    sheet = book.active
    # LoginPage is the name of the excel sheet in entire worksheet
    if "LoginPage" in book.sheetnames:
        sheet = book.worksheets[0]
        # second row first column
        #     username = sheet.cell(row=2,column=1).value
        username = sheet.cell(row=rownum, column=columnnum).value

    return username


print(read_username_data(2, 1))
